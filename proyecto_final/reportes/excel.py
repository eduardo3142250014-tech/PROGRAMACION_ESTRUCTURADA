from openpyxl import Workbook
from openpyxl.styles import Font

from reportes import datos
import funciones


def generar():

    funciones.titulo("GENERAR REPORTE EXCEL")

    id_registro = funciones.leerEntero(
        "Ingrese el ID del registro: "
    )

    reporte = datos.obtenerReporte(
        id_registro
    )

    if reporte is None:

        print("\nNo existe ese registro.")

        funciones.espereTecla()

        return

    registro = reporte["registro"]

    cuerdas = reporte["cuerdas"]

    libro = Workbook()

    hoja = libro.active

    hoja.title = "Reporte"

    hoja["A1"] = "REPORTE DEL SISTEMA DE TENSIONES"

    hoja["A1"].font = Font(
        bold=True,
        size=16
    )

    fila = 3

    hoja[f"A{fila}"] = "ID Registro"
    hoja[f"B{fila}"] = registro[0]

    fila += 1

    hoja[f"A{fila}"] = "Proyecto"
    hoja[f"B{fila}"] = registro[1]

    fila += 1

    hoja[f"A{fila}"] = "Descripción"
    hoja[f"B{fila}"] = registro[2]

    fila += 1

    hoja[f"A{fila}"] = "Peso (kg)"
    hoja[f"B{fila}"] = registro[3]

    fila += 1

    hoja[f"A{fila}"] = "Posición"

    hoja[f"B{fila}"] = (

        f"({registro[4]}, "

        f"{registro[5]}, "

        f"{registro[6]})"

    )

    fila += 1

    hoja[f"A{fila}"] = "Número de cuerdas"
    hoja[f"B{fila}"] = registro[7]

    fila += 1

    hoja[f"A{fila}"] = "Resultado"
    hoja[f"B{fila}"] = registro[8]

    fila += 1

    hoja[f"A{fila}"] = "Factor de Seguridad"
    hoja[f"B{fila}"] = round(
        registro[9],
        2
    )

    fila += 1

    hoja[f"A{fila}"] = "Viable"

    hoja[f"B{fila}"] = (

        "SI"

        if registro[10]

        else

        "NO"

    )

    fila += 1

    hoja[f"A{fila}"] = "Fecha"
    hoja[f"B{fila}"] = str(registro[11])

    fila += 1

    hoja[f"A{fila}"] = "Empleado"
    hoja[f"B{fila}"] = registro[12]

    fila += 3

    hoja[f"A{fila}"] = "CUERDAS"

    hoja[f"A{fila}"].font = Font(
        bold=True
    )

    fila += 1

    encabezados = [

        "Nombre",

        "Material",

        "Longitud",

        "Ángulo X",

        "Ángulo Y",

        "Ángulo Z",

        "Tensión",

        "Resistencia",

        "Factor Seguridad",

        "Estado"

    ]

    columna = 1

    for encabezado in encabezados:

        celda = hoja.cell(

            row=fila,

            column=columna

        )

        celda.value = encabezado

        celda.font = Font(
            bold=True
        )

        columna += 1

    fila += 1

    for cuerda in cuerdas:

        fs = 0

        if cuerda[6] != 0:

            fs = cuerda[7] / cuerda[6]

        estado = (

            "CORRECTO"

            if fs >= 3

            else

            "NO VIABLE"

        )

        hoja.cell(
            row=fila,
            column=1
        ).value = cuerda[0]

        hoja.cell(
            row=fila,
            column=2
        ).value = cuerda[1]

        hoja.cell(
            row=fila,
            column=3
        ).value = round(
            cuerda[2],
            3
        )

        hoja.cell(
            row=fila,
            column=4
        ).value = round(
            cuerda[3],
            2
        )

        hoja.cell(
            row=fila,
            column=5
        ).value = round(
            cuerda[4],
            2
        )

        hoja.cell(
            row=fila,
            column=6
        ).value = round(
            cuerda[5],
            2
        )

        hoja.cell(
            row=fila,
            column=7
        ).value = round(
            cuerda[6],
            2
        )

        hoja.cell(
            row=fila,
            column=8
        ).value = round(
            cuerda[7],
            2
        )

        hoja.cell(
            row=fila,
            column=9
        ).value = round(
            fs,
            2
        )

        hoja.cell(
            row=fila,
            column=10
        ).value = estado

        fila += 1

    for columna in hoja.columns:

        longitud = 0

        letra = columna[0].column_letter

        for celda in columna:

            if celda.value is not None:

                longitud = max(

                    longitud,

                    len(

                        str(

                            celda.value

                        )

                    )

                )

        hoja.column_dimensions[

            letra

        ].width = longitud + 3

    libro.save(

        f"Reporte_{id_registro}.xlsx"

    )

    print(

        "\nReporte Excel generado correctamente."

    )

    print(

        f"\nArchivo: Reporte_{id_registro}.xlsx"

    )

    funciones.espereTecla()